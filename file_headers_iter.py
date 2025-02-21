import openpyxl

# Global dictionary to store headers from each sheet
global_headers = {}

def extract_headers(file_path):
    """Extract headers from all sheets in an Excel file by detecting bold formatting, regardless of row position."""
    global global_headers
    global_headers.clear()  # Clear previous data

    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        headers = []
        found_headers = False  # Flag to stop after finding the first header row

        # Iterate through all rows and columns
        for row in sheet.iter_rows():
            row_headers = []  # Temporary storage for possible headers
            bold_count = 0

            for cell in row:
                if cell.value and cell.font and cell.font.bold:  # Detect bold cells
                    bold_count += 1
                    row_headers.append(cell.value)
                elif cell.value:
                    row_headers.append(cell.value)
                else:
                    row_headers.append("")  # Keep column structure

            # If at least 50% of the row is bold, consider it as the header row
            if bold_count >= (len(row_headers) * 0.5):
                headers = row_headers
                found_headers = True
                break  # Stop searching once headers are found

        if found_headers:
            global_headers[sheet_name] = headers

    return global_headers

# Call function to populate global_headers
file_path = "C:/Users/RAMESH SINGH/PycharmProjects/File_validation/Test_files/20250114-PNC_Original.xlsx"
extract_headers(file_path)

# Print the extracted headers without extra quotes
print("Extracted Headers:")
for sheet, headers in global_headers.items():
    print(f"\nSheet: {sheet}")
    print("Headers:", ', '.join(filter(None, headers)))  # Removes empty columns from display
